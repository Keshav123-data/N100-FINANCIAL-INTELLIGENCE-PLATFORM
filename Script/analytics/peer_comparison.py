# ============================================================
# SPRINT 3 — DAY 20
# PEER COMPARISON EXCEL REPORT
#
# N100 FINANCIAL INTELLIGENCE PLATFORM
#
# OUTPUT:
# output/peer_comparison.xlsx
#
# REQUIREMENTS:
# - 11 peer-group sheets
# - company_id
# - company_name
# - 20 metric columns
# - percentile rank for each metric
# - percentile colour coding
# - benchmark company highlighted
# - peer-group median row
# ============================================================

from pathlib import Path
import sqlite3
import sys
import traceback

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = (
    PROJECT_ROOT
    / "DB"
    / "nifty100.db"
)

PROCESSED_DATA = (
    PROJECT_ROOT
    / "Data"
    / "processed"
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)

OUTPUT_FILE = (
    OUTPUT_DIR
    / "peer_comparison.xlsx"
)


# ============================================================
# 10 REQUIRED PEER METRICS
# ============================================================

METRICS = [
    "ROE",
    "ROCE",
    "NPM",
    "D/E",
    "FCF",
    "PAT CAGR 5Y",
    "Revenue CAGR 5Y",
    "EPS CAGR 5Y",
    "Interest Coverage",
    "Asset Turnover",
]


# ============================================================
# EXCEL COLOURS
# ============================================================

GREEN_FILL = PatternFill(
    fill_type="solid",
    fgColor="C6EFCE"
)

YELLOW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFEB9C"
)

RED_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)

GOLD_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFD966"
)

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78"
)

MEDIAN_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7"
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True
)

BOLD_FONT = Font(
    bold=True
)

THIN_SIDE = Side(
    style="thin",
    color="D9E1F2"
)

BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE
)


# ============================================================
# ENGINE
# ============================================================

class PeerComparisonEngine:

    def __init__(self):

        self.database_path = DATABASE_PATH

        self.processed_data = PROCESSED_DATA

        self.output_dir = OUTPUT_DIR

        self.output_file = OUTPUT_FILE

        self.peer_percentiles = None

        self.peer_groups = None

        self.companies = None

        self.financial_ratios = None

        self.report_data = None

        self.benchmark_companies = {}

        self.output_dir.mkdir(
            parents=True,
            exist_ok=True
        )

    # ========================================================
    # LOAD DATABASE
    # ========================================================

    def load_database(self):

        print()
        print("=" * 70)
        print("DAY 20 — LOADING DATABASE")
        print("=" * 70)

        if not self.database_path.exists():

            raise FileNotFoundError(
                f"\nDatabase not found:\n"
                f"{self.database_path}"
            )

        print(
            f"Database:\n"
            f"{self.database_path}"
        )

        conn = sqlite3.connect(
            self.database_path
        )

        try:

            # ------------------------------------------------
            # List tables
            # ------------------------------------------------

            tables = pd.read_sql_query(
                """
                SELECT name
                FROM sqlite_master
                WHERE type='table'
                """,
                conn
            )

            table_names = set(
                tables["name"].tolist()
            )

            print()
            print("Tables found:")

            for table in sorted(
                table_names
            ):

                print(
                    f"  ✓ {table}"
                )

            # ------------------------------------------------
            # peer_percentiles
            # ------------------------------------------------

            if (
                "peer_percentiles"
                not in table_names
            ):

                raise RuntimeError(
                    "\npeer_percentiles table "
                    "not found.\n\n"
                    "Run Day 18 first:\n"
                    "python "
                    "Script\\analytics\\peer.py"
                )

            self.peer_percentiles = (
                pd.read_sql_query(
                    """
                    SELECT *
                    FROM peer_percentiles
                    """,
                    conn
                )
            )

            # ------------------------------------------------
            # peer_groups
            # ------------------------------------------------

            if (
                "peer_groups"
                in table_names
            ):

                self.peer_groups = (
                    pd.read_sql_query(
                        """
                        SELECT *
                        FROM peer_groups
                        """,
                        conn
                    )
                )

            # ------------------------------------------------
            # companies
            # ------------------------------------------------

            if (
                "companies"
                in table_names
            ):

                self.companies = (
                    pd.read_sql_query(
                        """
                        SELECT *
                        FROM companies
                        """,
                        conn
                    )
                )

            # ------------------------------------------------
            # financial ratios
            # ------------------------------------------------

            if (
                "financial_ratios"
                in table_names
            ):

                self.financial_ratios = (
                    pd.read_sql_query(
                        """
                        SELECT *
                        FROM financial_ratios
                        """,
                        conn
                    )
                )

        finally:

            conn.close()

        # ----------------------------------------------------
        # Clean columns
        # ----------------------------------------------------

        self.peer_percentiles.columns = (
            self.peer_percentiles
            .columns
            .astype(str)
            .str.strip()
        )

        if self.peer_groups is not None:

            self.peer_groups.columns = (
                self.peer_groups
                .columns
                .astype(str)
                .str.strip()
            )

        if self.companies is not None:

            self.companies.columns = (
                self.companies
                .columns
                .astype(str)
                .str.strip()
            )

        print()
        print(
            "Peer percentile rows:",
            len(self.peer_percentiles)
        )

        print(
            "Companies:",
            self.peer_percentiles[
                "company_id"
            ].nunique()
        )

        print(
            "Peer groups:",
            self.peer_percentiles[
                "peer_group_name"
            ].nunique()
        )

    # ========================================================
    # LOAD BENCHMARK COMPANIES
    # ========================================================

    def load_benchmarks(self):

        print()
        print("=" * 70)
        print("LOADING BENCHMARK COMPANIES")
        print("=" * 70)

        # ----------------------------------------------------
        # First try peer_groups table
        # ----------------------------------------------------

        if self.peer_groups is not None:

            columns = set(
                self.peer_groups.columns
            )

            benchmark_column = None

            possible_columns = [
                "is_benchmark",
                "benchmark",
                "benchmark_company",
                "is_benchmark_company"
            ]

            for column in possible_columns:

                if column in columns:

                    benchmark_column = column
                    break

            if benchmark_column is not None:

                benchmark_df = (
                    self.peer_groups.copy()
                )

                benchmark_values = (
                    benchmark_df[
                        benchmark_column
                    ]
                    .astype(str)
                    .str.lower()
                    .isin(
                        [
                            "1",
                            "true",
                            "yes",
                            "y"
                        ]
                    )
                )

                benchmark_df = (
                    benchmark_df[
                        benchmark_values
                    ]
                )

                for _, row in (
                    benchmark_df.iterrows()
                ):

                    peer_group = row[
                        "peer_group_name"
                    ]

                    company_id = row[
                        "company_id"
                    ]

                    self.benchmark_companies[
                        peer_group
                    ] = company_id

        # ----------------------------------------------------
        # If no benchmark metadata exists
        # ----------------------------------------------------
        #
        # We do NOT invent a benchmark company.
        # Instead, attempt common benchmark columns.
        # ----------------------------------------------------

        if not self.benchmark_companies:

            print(
                "⚠ No explicit benchmark "
                "column found."
            )

            print(
                "Benchmark highlighting will "
                "be skipped unless benchmark "
                "information exists in the data."
            )

        else:

            for group, company in (
                self.benchmark_companies.items()
            ):

                print(
                    f"✓ {group}: {company}"
                )

    # ========================================================
    # COMPANY NAME LOOKUP
    # ========================================================

    def add_company_names(
        self,
        df
    ):

        result = df.copy()

        # ----------------------------------------------------
        # If names already exist
        # ----------------------------------------------------

        if "company_name" in result.columns:

            result["company_name"] = (
                result["company_name"]
                .fillna(
                    result["company_id"]
                )
            )

            return result

        # ----------------------------------------------------
        # companies table
        # ----------------------------------------------------

        if self.companies is not None:

            companies = (
                self.companies.copy()
            )

            id_column = None

            if "company_id" in companies.columns:

                id_column = "company_id"

            elif "id" in companies.columns:

                id_column = "id"

            name_column = None

            possible_name_columns = [
                "company_name",
                "name",
                "Company Name"
            ]

            for column in possible_name_columns:

                if column in companies.columns:

                    name_column = column
                    break

            if (
                id_column is not None
                and
                name_column is not None
            ):

                lookup = companies[
                    [
                        id_column,
                        name_column
                    ]
                ].copy()

                lookup.columns = [
                    "company_id",
                    "company_name"
                ]

                result = result.merge(
                    lookup,
                    on="company_id",
                    how="left"
                )

        # ----------------------------------------------------
        # Fallback
        # ----------------------------------------------------

        if "company_name" not in result.columns:

            result["company_name"] = (
                result["company_id"]
            )

        result["company_name"] = (
            result["company_name"]
            .fillna(
                result["company_id"]
            )
            .astype(str)
            .str.strip()
        )

        return result

    # ========================================================
    # PREPARE REPORT DATA
    # ========================================================

    def prepare_report_data(self):

        print()
        print("=" * 70)
        print("PREPARING PEER COMPARISON DATA")
        print("=" * 70)

        df = self.peer_percentiles.copy()

        # ----------------------------------------------------
        # Required columns
        # ----------------------------------------------------

        required = {
            "company_id",
            "peer_group_name",
            "metric",
            "value",
            "percentile_rank"
        }

        missing = (
            required -
            set(df.columns)
        )

        if missing:

            raise ValueError(
                "Missing columns in "
                "peer_percentiles: "
                + ", ".join(
                    sorted(missing)
                )
            )

        # ----------------------------------------------------
        # Clean
        # ----------------------------------------------------

        df["company_id"] = (
            df["company_id"]
            .astype(str)
            .str.strip()
        )

        df["peer_group_name"] = (
            df["peer_group_name"]
            .astype(str)
            .str.strip()
        )

        df["metric"] = (
            df["metric"]
            .astype(str)
            .str.strip()
        )

        df["value"] = pd.to_numeric(
            df["value"],
            errors="coerce"
        )

        df["percentile_rank"] = pd.to_numeric(
            df["percentile_rank"],
            errors="coerce"
        )

        # ----------------------------------------------------
        # Only required 10 metrics
        # ----------------------------------------------------

        df = df[
            df["metric"].isin(
                METRICS
            )
        ].copy()

        # ----------------------------------------------------
        # Add company names
        # ----------------------------------------------------

        df = self.add_company_names(
            df
        )

        self.report_data = df

        print(
            f"Rows for report: "
            f"{len(df)}"
        )

        print(
            f"Peer groups: "
            f"{df['peer_group_name'].nunique()}"
        )

        print(
            f"Companies: "
            f"{df['company_id'].nunique()}"
        )

        # ----------------------------------------------------
        # Validate 11 peer groups
        # ----------------------------------------------------

        peer_groups = sorted(
            df[
                "peer_group_name"
            ]
            .dropna()
            .unique()
        )

        print()
        print(
            "Peer groups found:"
        )

        for group in peer_groups:

            count = (
                df[
                    df["peer_group_name"]
                    == group
                ]["company_id"]
                .nunique()
            )

            print(
                f"  ✓ {group}: "
                f"{count} companies"
            )

        if len(peer_groups) != 11:

            print()
            print(
                f"⚠ Expected 11 peer groups, "
                f"found {len(peer_groups)}"
            )

    # ========================================================
    # CREATE WIDE SHEET
    # ========================================================

    def create_peer_sheet(
        self,
        peer_group_name
    ):

        df = self.report_data[
            self.report_data[
                "peer_group_name"
            ]
            == peer_group_name
        ].copy()

        # ----------------------------------------------------
        # Values
        # ----------------------------------------------------

        values = df.pivot_table(
            index=[
                "company_id",
                "company_name"
            ],
            columns="metric",
            values="value",
            aggfunc="first"
        )

        # ----------------------------------------------------
        # Percentiles
        # ----------------------------------------------------

        percentiles = df.pivot_table(
            index=[
                "company_id",
                "company_name"
            ],
            columns="metric",
            values="percentile_rank",
            aggfunc="first"
        )

        # ----------------------------------------------------
        # Flatten
        # ----------------------------------------------------

        values = values.reset_index()

        percentiles = (
            percentiles
            .reset_index()
        )

        # ----------------------------------------------------
        # Ensure every metric exists
        # ----------------------------------------------------

        for metric in METRICS:

            if metric not in values.columns:

                values[metric] = np.nan

            if (
                metric
                not in percentiles.columns
            ):

                percentiles[metric] = np.nan

        # ----------------------------------------------------
        # Select order
        # ----------------------------------------------------

        values = values[
            [
                "company_id",
                "company_name"
            ]
            + METRICS
        ]

        percentiles = percentiles[
            [
                "company_id",
                "company_name"
            ]
            + METRICS
        ]

        # ----------------------------------------------------
        # Rename metric columns
        # ----------------------------------------------------

        value_columns = {
            metric:
                f"{metric} Value"
            for metric in METRICS
        }

        percentile_columns = {
            metric:
                f"{metric} Percentile"
            for metric in METRICS
        }

        values = values.rename(
            columns=value_columns
        )

        percentiles = percentiles.rename(
            columns=percentile_columns
        )

        # ----------------------------------------------------
        # Merge
        # ----------------------------------------------------

        result = values.merge(
            percentiles,
            on=[
                "company_id",
                "company_name"
            ],
            how="left"
        )

        # ----------------------------------------------------
        # Final column order
        # ----------------------------------------------------

        columns = [
            "company_id",
            "company_name"
        ]

        for metric in METRICS:

            columns.append(
                f"{metric} Value"
            )

            columns.append(
                f"{metric} Percentile"
            )

        result = result[
            columns
        ]

        # ----------------------------------------------------
        # Sort by average percentile
        # ----------------------------------------------------

        percentile_columns = [
            f"{metric} Percentile"
            for metric in METRICS
        ]

        result[
            "_average_percentile"
        ] = result[
            percentile_columns
        ].mean(
            axis=1,
            skipna=True
        )

        result = (
            result
            .sort_values(
                "_average_percentile",
                ascending=False
            )
            .drop(
                columns="_average_percentile"
            )
            .reset_index(
                drop=True
            )
        )

        return result

    # ========================================================
    # WRITE EXCEL
    # ========================================================

    def write_excel(self):

        print()
        print("=" * 70)
        print("CREATING EXCEL REPORT")
        print("=" * 70)

        peer_groups = sorted(
            self.report_data[
                "peer_group_name"
            ]
            .dropna()
            .unique()
        )

        # ----------------------------------------------------
        # ExcelWriter
        # ----------------------------------------------------

        with pd.ExcelWriter(
            self.output_file,
            engine="openpyxl"
        ) as writer:

            for peer_group in peer_groups:

                print(
                    f"\nCreating sheet: "
                    f"{peer_group}"
                )

                sheet_df = (
                    self.create_peer_sheet(
                        peer_group
                    )
                )

                # ------------------------------------------------
                # Excel sheet names max = 31 chars
                # ------------------------------------------------

                sheet_name = (
                    self.clean_sheet_name(
                        peer_group
                    )
                )

                # ------------------------------------------------
                # Write starting row 1
                # ------------------------------------------------

                sheet_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=0
                )

                print(
                    f"  ✓ "
                    f"{len(sheet_df)} companies"
                )

        print()
        print(
            f"Excel created:\n"
            f"{self.output_file}"
        )

    # ========================================================
    # CLEAN EXCEL SHEET NAME
    # ========================================================

    @staticmethod
    def clean_sheet_name(
        name
    ):

        name = str(name)

        invalid = [
            "\\",
            "/",
            "*",
            "?",
            ":",
            "[",
            "]"
        ]

        for char in invalid:

            name = name.replace(
                char,
                "_"
            )

        return name[:31]

    # ========================================================
    # FORMAT EXCEL
    # ========================================================

    def format_excel(self):

        print()
        print("=" * 70)
        print("FORMATTING EXCEL REPORT")
        print("=" * 70)

        if not self.output_file.exists():

            raise FileNotFoundError(
                "Excel output file was not created."
            )

        workbook = load_workbook(
            self.output_file
        )

        # ----------------------------------------------------
        # Each sheet
        # ----------------------------------------------------

        for worksheet in workbook.worksheets:

            print(
                f"Formatting: "
                f"{worksheet.title}"
            )

            # ------------------------------------------------
            # Header
            # ------------------------------------------------

            for cell in worksheet[1]:

                cell.fill = HEADER_FILL

                cell.font = WHITE_FONT

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                cell.border = BORDER

            worksheet.row_dimensions[
                1
            ].height = 30

            # ------------------------------------------------
            # Freeze top row
            # ------------------------------------------------

            worksheet.freeze_panes = "C2"

            # ------------------------------------------------
            # Autofilter
            # ------------------------------------------------

            worksheet.auto_filter.ref = (
                worksheet.dimensions
            )

            # ------------------------------------------------
            # Find columns
            # ------------------------------------------------

            headers = {}

            for cell in worksheet[1]:

                headers[
                    cell.value
                ] = cell.column

            # ------------------------------------------------
            # Format data cells
            # ------------------------------------------------

            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row
            ):

                for cell in row:

                    cell.border = BORDER

                    cell.alignment = Alignment(
                        vertical="center"
                    )

            # ------------------------------------------------
            # Percentile colour coding
            #
            # Green >= 75
            # Yellow 25–75
            # Red <= 25
            # ------------------------------------------------

            for metric in METRICS:

                column_name = (
                    f"{metric} Percentile"
                )

                if (
                    column_name
                    not in headers
                ):

                    continue

                column_number = (
                    headers[
                        column_name
                    ]
                )

                for row_number in range(
                    2,
                    worksheet.max_row + 1
                ):

                    cell = worksheet.cell(
                        row=row_number,
                        column=column_number
                    )

                    value = cell.value

                    if value is None:

                        continue

                    try:

                        value = float(
                            value
                        )

                    except (
                        TypeError,
                        ValueError
                    ):

                        continue

                    # ----------------------------------------
                    # Green
                    # ----------------------------------------

                    if value >= 75:

                        cell.fill = (
                            GREEN_FILL
                        )

                    # ----------------------------------------
                    # Yellow
                    # ----------------------------------------

                    elif value > 25:

                        cell.fill = (
                            YELLOW_FILL
                        )

                    # ----------------------------------------
                    # Red
                    # ----------------------------------------

                    else:

                        cell.fill = (
                            RED_FILL
                        )

                    cell.alignment = (
                        Alignment(
                            horizontal="center"
                        )
                    )

            # ------------------------------------------------
            # Benchmark row
            # ------------------------------------------------

            benchmark_id = (
                self.benchmark_companies
                .get(
                    worksheet.title
                )
            )

            # If cleaned sheet name doesn't match,
            # find original peer group.
            if benchmark_id is None:

                for peer_group, company_id in (
                    self.benchmark_companies.items()
                ):

                    if (
                        self.clean_sheet_name(
                            peer_group
                        )
                        ==
                        worksheet.title
                    ):

                        benchmark_id = company_id

                        break

            if benchmark_id is not None:

                for row_number in range(
                    2,
                    worksheet.max_row + 1
                ):

                    company_id = (
                        worksheet.cell(
                            row=row_number,
                            column=1
                        ).value
                    )

                    if (
                        str(company_id).strip()
                        ==
                        str(benchmark_id).strip()
                    ):

                        for cell in (
                            worksheet[row_number]
                        ):

                            cell.fill = (
                                GOLD_FILL
                            )

                            cell.font = (
                                BOLD_FONT
                            )

                        print(
                            f"  ✓ Benchmark "
                            f"highlighted: "
                            f"{benchmark_id}"
                        )

                        break

            # ------------------------------------------------
            # Median row
            # ------------------------------------------------

            median_row = (
                worksheet.max_row + 2
            )

            worksheet.cell(
                row=median_row,
                column=1,
                value="PEER GROUP MEDIAN"
            )

            worksheet.cell(
                row=median_row,
                column=1
            ).fill = MEDIAN_FILL

            worksheet.cell(
                row=median_row,
                column=1
            ).font = BOLD_FONT

            worksheet.cell(
                row=median_row,
                column=2,
                value=""
            )

            worksheet.cell(
                row=median_row,
                column=2
            ).fill = MEDIAN_FILL

            # ----------------------------------------------
            # Median for numeric columns
            # ----------------------------------------------

            for column_number in range(
                3,
                worksheet.max_column + 1
            ):

                values = []

                for row_number in range(
                    2,
                    worksheet.max_row + 1
                ):

                    value = worksheet.cell(
                        row=row_number,
                        column=column_number
                    ).value

                    if isinstance(
                        value,
                        (int, float)
                    ):

                        if np.isfinite(
                            value
                        ):

                            values.append(
                                value
                            )

                if values:

                    median = float(
                        np.median(values)
                    )

                    worksheet.cell(
                        row=median_row,
                        column=column_number,
                        value=round(
                            median,
                            2
                        )
                    )

                worksheet.cell(
                    row=median_row,
                    column=column_number
                ).fill = MEDIAN_FILL

                worksheet.cell(
                    row=median_row,
                    column=column_number
                ).font = BOLD_FONT

                worksheet.cell(
                    row=median_row,
                    column=column_number
                ).border = BORDER

            # ------------------------------------------------
            # Median row styling
            # ------------------------------------------------

            for cell in worksheet[
                median_row
            ]:

                cell.fill = MEDIAN_FILL

                cell.font = BOLD_FONT

                cell.border = BORDER

                cell.alignment = Alignment(
                    horizontal="center"
                )

            # ------------------------------------------------
            # Column widths
            # ------------------------------------------------

            for column_cells in (
                worksheet.columns
            ):

                column_letter = (
                    get_column_letter(
                        column_cells[0].column
                    )
                )

                max_length = 0

                for cell in column_cells:

                    try:

                        value_length = len(
                            str(cell.value)
                        )

                        if (
                            value_length
                            > max_length
                        ):

                            max_length = (
                                value_length
                            )

                    except Exception:

                        pass

                width = min(
                    max(
                        max_length + 2,
                        12
                    ),
                    30
                )

                worksheet.column_dimensions[
                    column_letter
                ].width = width

            # ------------------------------------------------
            # Number formatting
            # ------------------------------------------------

            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row
            ):

                for cell in row:

                    if isinstance(
                        cell.value,
                        (int, float)
                    ):

                        cell.number_format = (
                            "0.00"
                        )

            # ------------------------------------------------
            # Page setup
            # ------------------------------------------------

            worksheet.sheet_view.showGridLines = (
                False
            )

            worksheet.page_setup.orientation = (
                "landscape"
            )

            worksheet.page_setup.fitToWidth = 1

            worksheet.page_setup.fitToHeight = 0

            worksheet.sheet_properties.pageSetUpPr.fitToPage = (
                True
            )

        # ----------------------------------------------------
        # Save
        # ----------------------------------------------------

        workbook.save(
            self.output_file
        )

        print()
        print(
            "✓ Excel formatting completed"
        )

    # ========================================================
    # VALIDATE REPORT
    # ========================================================

    def validate_report(self):

        print()
        print("=" * 70)
        print("DAY 20 VALIDATION")
        print("=" * 70)

        if not self.output_file.exists():

            raise RuntimeError(
                "peer_comparison.xlsx "
                "was not generated."
            )

        workbook = load_workbook(
            self.output_file,
            read_only=True
        )

        sheet_count = len(
            workbook.sheetnames
        )

        print(
            f"Excel sheets: "
            f"{sheet_count}"
        )

        print()
        print("Sheets:")

        for sheet in workbook.sheetnames:

            print(
                f"  ✓ {sheet}"
            )

        # ----------------------------------------------------
        # Requirement: exactly 11 sheets
        # ----------------------------------------------------

        if sheet_count != 11:

            print()
            print(
                "⚠ WARNING:"
            )

            print(
                f"Expected 11 sheets, "
                f"found {sheet_count}"
            )

        else:

            print()
            print(
                "✓ Exactly 11 peer-group "
                "sheets found"
            )

        # ----------------------------------------------------
        # Check columns
        # ----------------------------------------------------

        for sheet_name in (
            workbook.sheetnames
        ):

            worksheet = workbook[
                sheet_name
            ]

            headers = [
                cell.value
                for cell in worksheet[1]
            ]

            if (
                "company_id"
                not in headers
            ):

                raise RuntimeError(
                    f"{sheet_name}: "
                    "company_id missing"
                )

            if (
                "company_name"
                not in headers
            ):

                raise RuntimeError(
                    f"{sheet_name}: "
                    "company_name missing"
                )

            for metric in METRICS:

                value_column = (
                    f"{metric} Value"
                )

                percentile_column = (
                    f"{metric} Percentile"
                )

                if (
                    value_column
                    not in headers
                ):

                    raise RuntimeError(
                        f"{sheet_name}: "
                        f"{value_column} missing"
                    )

                if (
                    percentile_column
                    not in headers
                ):

                    raise RuntimeError(
                        f"{sheet_name}: "
                        f"{percentile_column} missing"
                    )

        workbook.close()

        print(
            "✓ Required columns verified"
        )

        print(
            "✓ Day 20 report validation "
            "completed"
        )


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    print()
    print("=" * 70)
    print("SPRINT 3 — DAY 20")
    print("PEER COMPARISON EXCEL REPORT")
    print("=" * 70)

    try:

        engine = PeerComparisonEngine()

        # Step 1
        engine.load_database()

        # Step 2
        engine.load_benchmarks()

        # Step 3
        engine.prepare_report_data()

        # Step 4
        engine.write_excel()

        # Step 5
        engine.format_excel()

        # Step 6
        engine.validate_report()

        print()
        print("=" * 70)
        print("DAY 20 COMPLETED SUCCESSFULLY")
        print("=" * 70)

        print()
        print(
            "Output:"
        )

        print(
            engine.output_file
        )

    except Exception as e:

        print()
        print("=" * 70)
        print("ERROR")
        print("=" * 70)

        print(
            f"{type(e).__name__}: {e}"
        )

        print()

        traceback.print_exc()

        sys.exit(1)