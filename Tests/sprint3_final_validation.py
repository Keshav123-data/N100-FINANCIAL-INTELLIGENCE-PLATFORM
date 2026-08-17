from pathlib import Path
import sqlite3
import sys

import pandas as pd
from openpyxl import load_workbook


# ============================================================
# PROJECT PATH
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DB_PATH = PROJECT_ROOT / "DB" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"

FINANCIAL_RATIOS = OUTPUT_DIR / "financial_ratios.csv"
SCREENER_OUTPUT = OUTPUT_DIR / "screener_output.xlsx"
PEER_OUTPUT = OUTPUT_DIR / "peer_comparison.xlsx"


# ============================================================
# HELPERS
# ============================================================

passed = 0
failed = 0


def check(condition, message):
    global passed, failed

    if condition:
        print(f"  ✓ {message}")
        passed += 1
    else:
        print(f"  ❌ {message}")
        failed += 1


# ============================================================
# MAIN VALIDATION
# ============================================================

print()
print("=" * 80)
print("SPRINT 3 — DAY 21")
print("FINAL INTEGRATION & VALIDATION")
print("=" * 80)


# ============================================================
# 1. CHECK PROJECT OUTPUT FILES
# ============================================================

print()
print("=" * 80)
print("1. OUTPUT FILE VALIDATION")
print("=" * 80)

check(
    DB_PATH.exists(),
    "SQLite database exists"
)

check(
    FINANCIAL_RATIOS.exists(),
    "financial_ratios.csv exists"
)

check(
    SCREENER_OUTPUT.exists(),
    "screener_output.xlsx exists"
)

check(
    PEER_OUTPUT.exists(),
    "peer_comparison.xlsx exists"
)


# ============================================================
# 2. FINANCIAL RATIOS VALIDATION
# ============================================================

print()
print("=" * 80)
print("2. FINANCIAL RATIOS VALIDATION")
print("=" * 80)

if FINANCIAL_RATIOS.exists():

    ratios = pd.read_csv(
        FINANCIAL_RATIOS
    )

    print(
        f"Rows: {len(ratios)}"
    )

    print(
        f"Columns: {len(ratios.columns)}"
    )

    check(
        len(ratios) > 0,
        "Financial ratios contains records"
    )

    check(
        "company_id" in ratios.columns,
        "company_id column exists"
    )

    check(
        "year" in ratios.columns,
        "year column exists"
    )

    required_ratio_columns = [
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "return_on_equity_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "revenue_cagr_5yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "composite_quality_score",
    ]

    for column in required_ratio_columns:

        check(
            column in ratios.columns,
            f"{column} exists"
        )

    duplicates = (
        ratios
        .groupby(
            ["company_id", "year"]
        )
        .size()
    )

    duplicate_count = (
        duplicates > 1
    ).sum()

    check(
        duplicate_count == 0,
        "No duplicate company/year records"
    )


# ============================================================
# 3. SCREENER VALIDATION
# ============================================================

print()
print("=" * 80)
print("3. SCREENER OUTPUT VALIDATION")
print("=" * 80)

if SCREENER_OUTPUT.exists():

    workbook = load_workbook(
        SCREENER_OUTPUT,
        read_only=True
    )

    sheets = workbook.sheetnames

    print(
        f"Sheets: {len(sheets)}"
    )

    expected_screener_sheets = [
        "Quality Compounder",
        "Value Pick",
        "Growth Accelerator",
        "Dividend Champion",
        "Debt-Free Blue Chip",
        "Turnaround Watch",
    ]

    for sheet in expected_screener_sheets:

        check(
            sheet in sheets,
            f"Screener sheet exists: {sheet}"
        )

    check(
        len(sheets) == 6,
        "Exactly 6 screener preset sheets"
    )

    workbook.close()


# ============================================================
# 4. PEER COMPARISON VALIDATION
# ============================================================

print()
print("=" * 80)
print("4. PEER COMPARISON VALIDATION")
print("=" * 80)

if PEER_OUTPUT.exists():

    workbook = load_workbook(
        PEER_OUTPUT,
        read_only=True
    )

    peer_sheets = workbook.sheetnames

    print(
        f"Peer sheets: {len(peer_sheets)}"
    )

    check(
        len(peer_sheets) == 11,
        "Exactly 11 peer-group sheets"
    )

    for sheet in peer_sheets:

        worksheet = workbook[
            sheet
        ]

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        check(
            "company_id" in headers,
            f"{sheet}: company_id exists"
        )

        check(
            "company_name" in headers,
            f"{sheet}: company_name exists"
        )

        for metric in [
            "ROE",
            "ROCE",
            "NPM",
            "D/E",
            "FCF",
            "PAT CAGR 5Y",
            "Revenue CAGR 5Y",
            "EPS CAGR 5Y",
            "Interest Coverage",
            "Asset Turnover",
        ]:

            check(
                f"{metric} Value" in headers,
                f"{sheet}: {metric} Value exists"
            )

            check(
                f"{metric} Percentile" in headers,
                f"{sheet}: {metric} Percentile exists"
            )

    workbook.close()


# ============================================================
# 5. SQLITE VALIDATION
# ============================================================

print()
print("=" * 80)
print("5. SQLITE DATABASE VALIDATION")
print("=" * 80)

if DB_PATH.exists():

    conn = sqlite3.connect(
        DB_PATH
    )

    try:

        tables = pd.read_sql_query(
            """
            SELECT name
            FROM sqlite_master
            WHERE type='table'
            """,
            conn
        )

        table_names = set(
            tables["name"]
        )

        required_tables = [
            "companies",
            "financial_ratios",
            "peer_groups",
            "peer_percentiles",
        ]

        for table in required_tables:

            check(
                table in table_names,
                f"SQLite table exists: {table}"
            )

        # -----------------------------------------------
        # Financial ratios count
        # -----------------------------------------------

        if "financial_ratios" in table_names:

            count = pd.read_sql_query(
                """
                SELECT COUNT(*) AS count
                FROM financial_ratios
                """,
                conn
            ).iloc[0]["count"]

            print(
                f"Financial ratio DB rows: {count}"
            )

            check(
                count > 0,
                "financial_ratios table contains data"
            )

        # -----------------------------------------------
        # Peer percentile count
        # -----------------------------------------------

        if "peer_percentiles" in table_names:

            count = pd.read_sql_query(
                """
                SELECT COUNT(*) AS count
                FROM peer_percentiles
                """,
                conn
            ).iloc[0]["count"]

            print(
                f"Peer percentile DB rows: {count}"
            )

            check(
                count > 0,
                "peer_percentiles table contains data"
            )

    finally:

        conn.close()


# ============================================================
# 6. FINAL RESULT
# ============================================================

print()
print("=" * 80)
print("FINAL SPRINT 3 RESULT")
print("=" * 80)

print(
    f"Passed: {passed}"
)

print(
    f"Failed: {failed}"
)

print()

if failed == 0:

    print(
        "🎉 SPRINT 3 COMPLETED SUCCESSFULLY"
    )

    print()
    print(
        "All major Sprint 3 components are ready:"
    )

    print(
        "  ✓ Financial Ratio Engine"
    )

    print(
        "  ✓ Screener Engine"
    )

    print(
        "  ✓ Screener Presets"
    )

    print(
        "  ✓ Peer Percentile Engine"
    )

    print(
        "  ✓ Peer Comparison Excel"
    )

    print(
        "  ✓ SQLite Integration"
    )

    print(
        "  ✓ Final Validation"
    )

    print()
    print(
        "READY FOR SPRINT 4 — DASHBOARD & VALUATION"
    )

    sys.exit(0)

else:

    print(
        "❌ SPRINT 3 VALIDATION FAILED"
    )

    print(
        "Fix the failed checks before moving to Sprint 4."
    )

    sys.exit(1)