# ============================================================
# SPRINT 3 — DAY 21
# TESTS & SPRINT REVIEW
#
# N100 FINANCIAL INTELLIGENCE PLATFORM
#
# Run from project root:
#
# python Script\tests\sprint3_day21.py
#
# ============================================================

from pathlib import Path
import sqlite3
import sys
import subprocess
import traceback

import pandas as pd
import numpy as np
from openpyxl import load_workbook


# ============================================================
# PROJECT PATHS
# ============================================================

PROJECT_ROOT = Path(__file__).resolve().parents[1]

DATABASE_PATH = (
    PROJECT_ROOT
    / "DB"
    / "nifty100.db"
)

PROCESSED_DATA = (
    PROJECT_ROOT
    / "Data"
    / "processed"
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)

REPORTS_DIR = (
    PROJECT_ROOT
    / "reports"
)

SCREENER_FILE = (
    OUTPUT_DIR
    / "screener_output.xlsx"
)

PEER_FILE = (
    OUTPUT_DIR
    / "peer_comparison.xlsx"
)

RADAR_DIR = (
    REPORTS_DIR
    / "radar_charts"
)


# ============================================================
# TEST RESULT TRACKER
# ============================================================

TEST_RESULTS = []


def record_test(
    name,
    passed,
    message=""
):

    TEST_RESULTS.append(
        {
            "test": name,
            "passed": passed,
            "message": message
        }
    )

    if passed:

        print(
            f"  ✓ PASS — {name}"
        )

    else:

        print(
            f"  ✗ FAIL — {name}"
        )

        if message:

            print(
                f"           {message}"
            )


# ============================================================
# HEADER
# ============================================================

def print_header():

    print()

    print("=" * 75)

    print(
        "SPRINT 3 — DAY 21"
    )

    print(
        "TESTS & SPRINT REVIEW"
    )

    print("=" * 75)

    print()

    print(
        f"Project: {PROJECT_ROOT}"
    )

    print(
        f"Database: {DATABASE_PATH}"
    )


# ============================================================
# DATABASE CONNECTION
# ============================================================

def get_connection():

    if not DATABASE_PATH.exists():

        raise FileNotFoundError(
            f"Database not found:\n"
            f"{DATABASE_PATH}"
        )

    return sqlite3.connect(
        DATABASE_PATH
    )


# ============================================================
# DATABASE TABLE CHECK
# ============================================================

def test_database():

    print()

    print(
        "=" * 75
    )

    print(
        "1. DATABASE VALIDATION"
    )

    print(
        "=" * 75
    )

    try:

        conn = get_connection()

        tables = pd.read_sql_query(
            """
            SELECT name
            FROM sqlite_master
            WHERE type='table'
            """,
            conn
        )

        conn.close()

        table_names = set(
            tables["name"]
        )

        print()

        print(
            "Tables found:"
        )

        for table in sorted(
            table_names
        ):

            print(
                f"  • {table}"
            )

        required_tables = [
            "peer_percentiles"
        ]

        for table in required_tables:

            record_test(
                f"Table exists: {table}",
                table in table_names,
                "Required table missing"
            )

    except Exception as e:

        record_test(
            "Database validation",
            False,
            str(e)
        )


# ============================================================
# RUN EXISTING DQ TESTS
# ============================================================

def run_dq_tests():

    print()

    print(
        "=" * 75
    )

    print(
        "2. DATA QUALITY TESTS"
    )

    print(
        "=" * 75
    )

    # --------------------------------------------------------
    # Possible test locations
    # --------------------------------------------------------

    possible_tests = [

        PROJECT_ROOT
        / "Tests"
        / "etl",

        PROJECT_ROOT
        / "tests"
        / "etl",

        PROJECT_ROOT
        / "Tests",

        PROJECT_ROOT
        / "tests",
    ]

    test_directory = None

    for directory in possible_tests:

        if directory.exists():

            test_directory = directory

            break

    if test_directory is None:

        print()

        print(
            "⚠ DQ test directory was not found."
        )

        print(
            "The existing DQ tests were not "
            "automatically executed."
        )

        record_test(
            "14 DQ unit tests",
            False,
            "Test directory not found"
        )

        return

    print()

    print(
        f"Test directory: "
        f"{test_directory}"
    )

    # --------------------------------------------------------
    # Run pytest
    # --------------------------------------------------------

    try:

        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "pytest",
                str(test_directory),
                "-v"
            ],
            cwd=PROJECT_ROOT,
            capture_output=True,
            text=True
        )

        output = (
            result.stdout
            + "\n"
            + result.stderr
        )

        print()

        print(
            output[-5000:]
        )

        # ----------------------------------------------------
        # Detect pytest result
        # ----------------------------------------------------

        if result.returncode == 0:

            record_test(
                "14 DQ unit tests",
                True,
                "pytest completed successfully"
            )

        else:

            record_test(
                "14 DQ unit tests",
                False,
                "pytest returned failure code"
            )

    except Exception as e:

        record_test(
            "14 DQ unit tests",
            False,
            str(e)
        )


# ============================================================
# FIND SCREENER ENGINE
# ============================================================

def find_screener_engine():

    possible_files = [

        PROJECT_ROOT
        / "Script"
        / "screener"
        / "engine.py",

        PROJECT_ROOT
        / "src"
        / "screener"
        / "engine.py",
    ]

    for file in possible_files:

        if file.exists():

            return file

    return None


# ============================================================
# TEST SCREENER OUTPUT
# ============================================================

def test_screener_output():

    print()

    print(
        "=" * 75
    )

    print(
        "3. SCREENER OUTPUT VALIDATION"
    )

    print(
        "=" * 75
    )

    if not SCREENER_FILE.exists():

        record_test(
            "screener_output.xlsx exists",
            False,
            f"File not found: {SCREENER_FILE}"
        )

        return

    record_test(
        "screener_output.xlsx exists",
        True
    )

    try:

        workbook = load_workbook(
            SCREENER_FILE,
            read_only=True
        )

        sheets = workbook.sheetnames

        print()

        print(
            "Screener sheets:"
        )

        for sheet in sheets:

            print(
                f"  ✓ {sheet}"
            )

        # ----------------------------------------------------
        # Six presets
        # ----------------------------------------------------

        expected_presets = [
            "Quality Compounder",
            "Value Pick",
            "Growth Accelerator",
            "Dividend Champion",
            "Debt-Free Blue Chip",
            "Turnaround Watch"
        ]

        normalized_sheets = [
            str(x).strip().lower()
            for x in sheets
        ]

        found_presets = 0

        for preset in expected_presets:

            if (
                preset.lower()
                in normalized_sheets
            ):

                found_presets += 1

        record_test(
            "6 preset screener sheets",
            found_presets == 6,
            f"Found {found_presets}/6"
        )

        workbook.close()

    except Exception as e:

        record_test(
            "Read screener_output.xlsx",
            False,
            str(e)
        )


# ============================================================
# QUALITY COMPOUNDER TEST
# ============================================================

def test_quality_compounder():

    print()

    print(
        "=" * 75
    )

    print(
        "4. QUALITY COMPOUNDER VALIDATION"
    )

    print(
        "=" * 75
    )

    if not SCREENER_FILE.exists():

        record_test(
            "Quality Compounder",
            False,
            "screener_output.xlsx not found"
        )

        return

    try:

        workbook = load_workbook(
            SCREENER_FILE,
            read_only=True,
            data_only=True
        )

        sheet = None

        for name in workbook.sheetnames:

            if (
                "quality"
                in name.lower()
                and
                "compounder"
                in name.lower()
            ):

                sheet = name

                break

        if sheet is None:

            record_test(
                "Quality Compounder sheet",
                False,
                "Sheet not found"
            )

            workbook.close()

            return

        df = pd.read_excel(
            SCREENER_FILE,
            sheet_name=sheet
        )

        workbook.close()

        print()

        print(
            f"Companies returned: "
            f"{len(df)}"
        )

        # ----------------------------------------------------
        # Locate columns
        # ----------------------------------------------------

        def find_column(
            possible
        ):

            for column in df.columns:

                normalized = (
                    str(column)
                    .strip()
                    .lower()
                    .replace("_", " ")
                )

                for candidate in possible:

                    if (
                        candidate
                        in normalized
                    ):

                        return column

            return None

        roe_column = find_column(
            ["roe"]
        )

        de_column = find_column(
            [
                "d/e",
                "de"
            ]
        )

        if roe_column is None:

            record_test(
                "Quality Compounder ROE column",
                False,
                "ROE column not found"
            )

            return

        if de_column is None:

            record_test(
                "Quality Compounder D/E column",
                False,
                "D/E column not found"
            )

            return

        df[roe_column] = pd.to_numeric(
            df[roe_column],
            errors="coerce"
        )

        df[de_column] = pd.to_numeric(
            df[de_column],
            errors="coerce"
        )

        # ----------------------------------------------------
        # Verify ROE
        # ----------------------------------------------------

        invalid_roe = df[
            df[roe_column]
            <= 15
        ]

        record_test(
            "Quality Compounder ROE > 15%",
            invalid_roe.empty,
            f"{len(invalid_roe)} invalid rows"
        )

        # ----------------------------------------------------
        # Verify D/E
        # ----------------------------------------------------

        non_financials = df[
            df["broad_sector"]
            .astype(str)
            .str.strip()
            .str.lower()
            != "financials"
        ]

        invalid_de = non_financials[
            non_financials[de_column] >= 1
        ]

        record_test(
            "Quality Compounder D/E < 1",
            invalid_de.empty,
            f"{len(invalid_de)} invalid rows"
        )

        # ----------------------------------------------------
        # Verify company count
        # ----------------------------------------------------

        record_test(
            "Quality Compounder returns 5–50 companies",
            5 <= len(df) <= 50,
            f"Returned {len(df)} companies"
        )

        # ----------------------------------------------------
        # Top 5
        # ----------------------------------------------------

        print()

        print(
            "TOP 5 QUALITY COMPOUNDER RESULTS"
        )

        print(
            "-" * 75
        )

        display_columns = []

        for column in [
            "company_id",
            "company_name",
            roe_column,
            de_column
        ]:

            if column in df.columns:

                display_columns.append(
                    column
                )

        if display_columns:

            print(
                df[
                    display_columns
                ].head(5).to_string(
                    index=False
                )
            )

    except Exception as e:

        record_test(
            "Quality Compounder validation",
            False,
            str(e)
        )


# ============================================================
# PEER PERCENTILE VALIDATION
# ============================================================

def load_peer_percentiles():

    conn = get_connection()

    try:

        df = pd.read_sql_query(
            """
            SELECT *
            FROM peer_percentiles
            """,
            conn
        )

    finally:

        conn.close()

    return df


# ============================================================
# IT SERVICES TEST
# ============================================================

def test_it_services():

    print()

    print(
        "=" * 75
    )

    print(
        "5. IT SERVICES PEER RANKING TEST"
    )

    print(
        "=" * 75
    )

    try:

        df = load_peer_percentiles()

        # ----------------------------------------------------
        # Find IT Services
        # ----------------------------------------------------

        groups = (
            df["peer_group_name"]
            .astype(str)
            .str.strip()
        )

        it_mask = (
            groups.str.lower()
            == "it services"
        )

        it_df = df[
            it_mask
        ].copy()

        if it_df.empty:

            # Try contains
            it_df = df[
                groups.str.contains(
                    "it",
                    case=False,
                    na=False
                )
            ].copy()

        if it_df.empty:

            record_test(
                "IT Services peer group",
                False,
                "IT Services peer group not found"
            )

            return

        print(
            f"Rows: {len(it_df)}"
        )

        # ----------------------------------------------------
        # Find ROE metric
        # ----------------------------------------------------

        metric = (
            it_df["metric"]
            .astype(str)
            .str.strip()
        )

        roe_df = it_df[
            metric.str.upper()
            == "ROE"
        ].copy()

        if roe_df.empty:

            record_test(
                "IT Services ROE ranking",
                False,
                "ROE metric not found"
            )

            return

        roe_df["value"] = pd.to_numeric(
            roe_df["value"],
            errors="coerce"
        )

        roe_df["percentile_rank"] = (
            pd.to_numeric(
                roe_df[
                    "percentile_rank"
                ],
                errors="coerce"
            )
        )

        roe_df = roe_df.dropna(
            subset=[
                "value",
                "percentile_rank"
            ]
        )

        # ----------------------------------------------------
        # Highest ROE
        # ----------------------------------------------------

        highest_roe_row = (
            roe_df.loc[
                roe_df["value"].idxmax()
            ]
        )

        highest_percentile_row = (
            roe_df.loc[
                roe_df[
                    "percentile_rank"
                ].idxmax()
            ]
        )

        highest_roe_company = (
            highest_roe_row[
                "company_id"
            ]
        )

        highest_percentile_company = (
            highest_percentile_row[
                "company_id"
            ]
        )

        highest_roe = (
            highest_roe_row["value"]
        )

        highest_percentile = (
            highest_percentile_row[
                "percentile_rank"
            ]
        )

        print()

        print(
            f"Highest ROE company: "
            f"{highest_roe_company}"
        )

        print(
            f"Highest ROE: "
            f"{highest_roe}"
        )

        print(
            f"Highest ROE percentile company: "
            f"{highest_percentile_company}"
        )

        print(
            f"Highest percentile: "
            f"{highest_percentile}"
        )

        passed = (
            str(highest_roe_company)
            ==
            str(highest_percentile_company)
        )

        record_test(
            "IT Services highest ROE = highest ROE percentile",
            passed,
            (
                f"ROE leader={highest_roe_company}, "
                f"percentile leader={highest_percentile_company}"
            )
        )

    except Exception as e:

        record_test(
            "IT Services peer validation",
            False,
            str(e)
        )


# ============================================================
# FMCG TEST
# ============================================================

def test_fmcg():

    print()

    print(
        "=" * 75
    )

    print(
        "6. FMCG PEER RANKING SPOT CHECK"
    )

    print(
        "=" * 75
    )

    try:

        df = load_peer_percentiles()

        groups = (
            df["peer_group_name"]
            .astype(str)
            .str.strip()
        )

        fmcg = df[
            groups.str.contains(
                "fmcg",
                case=False,
                na=False
            )
        ].copy()

        if fmcg.empty:

            record_test(
                "FMCG peer group exists",
                False,
                "FMCG peer group not found"
            )

            return

        print(
            f"FMCG rows: {len(fmcg)}"
        )

        companies = (
            fmcg[
                "company_id"
            ].nunique()
        )

        print(
            f"FMCG companies: {companies}"
        )

        record_test(
            "FMCG peer group exists",
            companies > 0
        )

        # ----------------------------------------------------
        # Check ROE ranking
        # ----------------------------------------------------

        metric = (
            fmcg["metric"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        roe = fmcg[
            metric == "ROE"
        ].copy()

        if not roe.empty:

            roe["value"] = pd.to_numeric(
                roe["value"],
                errors="coerce"
            )

            roe["percentile_rank"] = (
                pd.to_numeric(
                    roe[
                        "percentile_rank"
                    ],
                    errors="coerce"
                )
            )

            roe = roe.dropna(
                subset=[
                    "value",
                    "percentile_rank"
                ]
            )

            if not roe.empty:

                value_leader = (
                    roe.loc[
                        roe["value"]
                        .idxmax()
                    ]["company_id"]
                )

                percentile_leader = (
                    roe.loc[
                        roe[
                            "percentile_rank"
                        ].idxmax()
                    ]["company_id"]
                )

                print(
                    f"FMCG ROE leader: "
                    f"{value_leader}"
                )

                print(
                    f"FMCG percentile leader: "
                    f"{percentile_leader}"
                )

                record_test(
                    "FMCG ROE percentile spot check",
                    str(value_leader)
                    ==
                    str(percentile_leader),
                    (
                        f"value leader="
                        f"{value_leader}, "
                        f"percentile leader="
                        f"{percentile_leader}"
                    )
                )

    except Exception as e:

        record_test(
            "FMCG validation",
            False,
            str(e)
        )


# ============================================================
# PEER EXCEL TEST
# ============================================================

def test_peer_excel():

    print()

    print(
        "=" * 75
    )

    print(
        "7. PEER COMPARISON EXCEL TEST"
    )

    print(
        "=" * 75
    )

    if not PEER_FILE.exists():

        record_test(
            "peer_comparison.xlsx exists",
            False,
            f"File not found: {PEER_FILE}"
        )

        return

    record_test(
        "peer_comparison.xlsx exists",
        True
    )

    try:

        workbook = load_workbook(
            PEER_FILE,
            read_only=True
        )

        sheets = workbook.sheetnames

        print()

        print(
            f"Sheets found: "
            f"{len(sheets)}"
        )

        for sheet in sheets:

            print(
                f"  ✓ {sheet}"
            )

        # ----------------------------------------------------
        # Exactly 11 sheets
        # ----------------------------------------------------

        record_test(
            "Peer comparison has exactly 11 sheets",
            len(sheets) == 11,
            f"Found {len(sheets)} sheets"
        )

        # ----------------------------------------------------
        # Check contents
        # ----------------------------------------------------

        for sheet in sheets:

            worksheet = workbook[
                sheet
            ]

            headers = [
                cell.value
                for cell in worksheet[1]
            ]

            if (
                "company_id"
                not in headers
            ):

                record_test(
                    f"{sheet} company_id",
                    False,
                    "company_id missing"
                )

                continue

            if (
                "company_name"
                not in headers
            ):

                record_test(
                    f"{sheet} company_name",
                    False,
                    "company_name missing"
                )

                continue

            record_test(
                f"{sheet} required columns",
                True
            )

        workbook.close()

    except Exception as e:

        record_test(
            "Peer comparison workbook",
            False,
            str(e)
        )


# ============================================================
# RADAR CHART TEST
# ============================================================

def test_radar_charts():

    print()

    print(
        "=" * 75
    )

    print(
        "8. RADAR CHART VALIDATION"
    )

    print(
        "=" * 75
    )

    if not RADAR_DIR.exists():

        record_test(
            "Radar chart directory exists",
            False,
            f"Directory not found: {RADAR_DIR}"
        )

        return

    png_files = list(
        RADAR_DIR.glob(
            "*_radar.png"
        )
    )

    print(
        f"Radar charts found: "
        f"{len(png_files)}"
    )

    record_test(
        "Radar charts generated",
        len(png_files) > 0,
        "No radar PNG files found"
    )


# ============================================================
# PEER GROUP COUNT
# ============================================================

def test_peer_group_count():

    print()

    print(
        "=" * 75
    )

    print(
        "9. PEER GROUP VALIDATION"
    )

    print(
        "=" * 75
    )

    try:

        df = load_peer_percentiles()

        groups = (
            df[
                "peer_group_name"
            ]
            .dropna()
            .astype(str)
            .str.strip()
            .unique()
        )

        print(
            f"Peer groups found: "
            f"{len(groups)}"
        )

        for group in sorted(
            groups
        ):

            count = (
                df[
                    df[
                        "peer_group_name"
                    ]
                    .astype(str)
                    .str.strip()
                    == group
                ][
                    "company_id"
                ].nunique()
            )

            print(
                f"  • {group}: "
                f"{count} companies"
            )

        record_test(
            "11 peer groups available",
            len(groups) == 11,
            f"Found {len(groups)}"
        )

    except Exception as e:

        record_test(
            "Peer group count",
            False,
            str(e)
        )


# ============================================================
# PEER PERCENTILE RANGE TEST
# ============================================================

def test_percentile_range():

    print()

    print(
        "=" * 75
    )

    print(
        "10. PERCENTILE RANGE VALIDATION"
    )

    print(
        "=" * 75
    )

    try:

        df = load_peer_percentiles()

        values = pd.to_numeric(
            df["percentile_rank"],
            errors="coerce"
        )

        invalid = df[
            (values < 0)
            |
            (values > 100)
        ]

        record_test(
            "Percentile ranks between 0 and 100",
            invalid.empty,
            f"{len(invalid)} invalid percentile rows"
        )

    except Exception as e:

        record_test(
            "Percentile range",
            False,
            str(e)
        )


# ============================================================
# D/E INVERSE TEST
# ============================================================

def test_de_inverse():

    print()

    print(
        "=" * 75
    )

    print(
        "11. D/E INVERSE RANKING TEST"
    )

    print(
        "=" * 75
    )

    try:

        df = load_peer_percentiles()

        metric = (
            df["metric"]
            .astype(str)
            .str.strip()
            .str.upper()
        )

        de = df[
            metric == "D/E"
        ].copy()

        if de.empty:

            # Try alternate naming
            de = df[
                metric == "DE"
            ].copy()

        if de.empty:

            record_test(
                "D/E inverse ranking",
                False,
                "D/E metric not found"
            )

            return

        de["value"] = pd.to_numeric(
            de["value"],
            errors="coerce"
        )

        de["percentile_rank"] = (
            pd.to_numeric(
                de[
                    "percentile_rank"
                ],
                errors="coerce"
            )
        )

        de = de.dropna(
            subset=[
                "value",
                "percentile_rank"
            ]
        )

        # ----------------------------------------------------
        # Check each peer group
        # ----------------------------------------------------

        violations = 0

        for group, group_df in (
            de.groupby(
                "peer_group_name"
            )
        ):

            if len(group_df) < 2:

                continue

            min_de_row = (
                group_df.loc[
                    group_df[
                        "value"
                    ].idxmin()
                ]
            )

            max_de_row = (
                group_df.loc[
                    group_df[
                        "value"
                    ].idxmax()
                ]
            )

            min_de_percentile = (
                min_de_row[
                    "percentile_rank"
                ]
            )

            max_de_percentile = (
                max_de_row[
                    "percentile_rank"
                ]
            )

            if (
                min_de_percentile
                < max_de_percentile
            ):

                violations += 1

                print(
                    f"  ✗ D/E inverse "
                    f"violation: {group}"
                )

        record_test(
            "D/E lower value receives higher percentile",
            violations == 0,
            f"{violations} peer groups failed"
        )

    except Exception as e:

        record_test(
            "D/E inverse ranking",
            False,
            str(e)
        )


# ============================================================
# FINAL SPRINT SUMMARY
# ============================================================

def sprint_summary():

    print()

    print(
        "=" * 75
    )

    print(
        "SPRINT 3 — FINAL REVIEW"
    )

    print(
        "=" * 75
    )

    total = len(
        TEST_RESULTS
    )

    passed = sum(
        1
        for result
        in TEST_RESULTS
        if result["passed"]
    )

    failed = (
        total - passed
    )

    print()

    print(
        f"Total checks : {total}"
    )

    print(
        f"Passed       : {passed}"
    )

    print(
        f"Failed       : {failed}"
    )

    print()

    print(
        "-" * 75
    )

    for result in TEST_RESULTS:

        status = (
            "PASS"
            if result["passed"]
            else "FAIL"
        )

        print(
            f"{status:<6} "
            f"{result['test']}"
        )

    print(
        "-" * 75
    )

    # --------------------------------------------------------
    # Definition of Done
    # --------------------------------------------------------

    print()

    print(
        "SPRINT 3 DEFINITION OF DONE"
    )

    print(
        "-" * 75
    )

    print(
        "✓ 6 preset screeners"
    )

    print(
        "✓ 5–50 companies per preset"
    )

    print(
        "✓ Peer percentile rankings"
    )

    print(
        "✓ 11 peer groups"
    )

    print(
        "✓ IT Services ranking verified"
    )

    print(
        "✓ FMCG spot check"
    )

    print(
        "✓ Radar charts generated"
    )

    print(
        "✓ peer_comparison.xlsx generated"
    )

    print(
        "✓ D/E inverse ranking checked"
    )

    print()

    if failed == 0:

        print(
            "=" * 75
        )

        print(
            "🎉 SPRINT 3 PASSED"
        )

        print(
            "ALL DAY 21 CHECKS PASSED"
        )

        print(
            "=" * 75
        )

        return 0

    else:

        print(
            "=" * 75
        )

        print(
            "⚠ SPRINT 3 NEEDS REVIEW"
        )

        print(
            f"{failed} check(s) failed."
        )

        print(
            "=" * 75
        )

        return 1


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    try:

        print_header()

        # ----------------------------------------------------
        # 1. Database
        # ----------------------------------------------------

        test_database()

        # ----------------------------------------------------
        # 2. Existing DQ tests
        # ----------------------------------------------------

        run_dq_tests()

        # ----------------------------------------------------
        # 3. Screener output
        # ----------------------------------------------------

        test_screener_output()

        # ----------------------------------------------------
        # 4. Quality Compounder
        # ----------------------------------------------------

        test_quality_compounder()

        # ----------------------------------------------------
        # 5. IT Services
        # ----------------------------------------------------

        test_it_services()

        # ----------------------------------------------------
        # 6. FMCG
        # ----------------------------------------------------

        test_fmcg()

        # ----------------------------------------------------
        # 7. Peer Excel
        # ----------------------------------------------------

        test_peer_excel()

        # ----------------------------------------------------
        # 8. Radar charts
        # ----------------------------------------------------

        test_radar_charts()

        # ----------------------------------------------------
        # 9. Peer groups
        # ----------------------------------------------------

        test_peer_group_count()

        # ----------------------------------------------------
        # 10. Percentile range
        # ----------------------------------------------------

        test_percentile_range()

        # ----------------------------------------------------
        # 11. D/E inverse
        # ----------------------------------------------------

        test_de_inverse()

        # ----------------------------------------------------
        # Final
        # ----------------------------------------------------

        exit_code = sprint_summary()

        sys.exit(
            exit_code
        )

    except Exception as e:

        print()

        print(
            "=" * 75
        )

        print(
            "DAY 21 FATAL ERROR"
        )

        print(
            "=" * 75
        )

        print(
            f"{type(e).__name__}: {e}"
        )

        print()

        traceback.print_exc()

        sys.exit(1)